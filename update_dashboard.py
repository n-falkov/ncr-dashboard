#!/usr/bin/env python3
"""
NCR Pilot Dashboard — Auto-updater
===================================
Reads GR xlsx + territory maps + activation stores,
generates index.html, pushes to GitHub Pages.

Usage:
  python3 update_dashboard.py

Expected files in data/:
  - GR_daily.xlsx or GR_weekly.xlsx  (General Report — daily or weekly data)
  - mgk_ath_territory_split_03_02*.xlsx  (Territory mapping Sep–Feb)
  - mgk_ath_territory_split_03_03*.xlsx  (Territory mapping Mar+, optional)
  - Overall_status_of_activation_stores.xlsx  (Activation stores)
"""

import os, sys, json, subprocess, re
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ─── CONFIG ──────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
DATA_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else SCRIPT_DIR / "data"
REPO_DIR = SCRIPT_DIR
OUTPUT_FILE = REPO_DIR / "index.html"
TEMPLATE_FILE = SCRIPT_DIR / "template.html"

# ─── TEAM CONFIGURATION (changes by month) ─────────────────────────
# Email aliases in March territory file
EMAIL_TO_NAME = {
    'a.hina@salmon.group': 'Aldrin Jhon Hina',
    'a.anonuevo@salmon.group': 'Arvin Anonuevo',
    'a.regonaos@salmon.group': 'Argenith Orlain Regonaos',
    'c.marasigan@salmon.group': 'Christy Marasigan',
    'd.navidad@salmon.group': 'Dad Ivan Natividad',
    'j.dastas@salmon.group': 'Jay Ar Dastas',
    'j.toledana@salmon.group': 'Jay Mark Toledana',
    'j.buac@salmon.group': 'Jhan-Jhan Buac',
    'm.mahinay@salmon.group': 'Ma. Cynthia Juano Mahinay',
    'm.melvida@salmon.group': 'Mary Anne Melvida',
    'o.casimo@salmon.group': 'Oscar Saverola Casimo',
    'r.sabala@salmon.group': 'Ricardo Sabala Jr.',
    'r.vargas@salmon.group': 'Roselyn Espinosa Vargas',
    't.nazareta@salmon.group': 'Tricxie Mae Nazareta',
    'h.sabarillo@salmon.group': 'Henry Sabarillo',
    'm.garcia1@salmon.group': 'Maria Sheena Garcia',
    'r.quito@salmon.group': 'Robin Mark Quito',
}

ATS_ALIASES = {
    "Mery Ann Melvida": "Mary Anne Melvida",
}
# Add email aliases
ATS_ALIASES.update(EMAIL_TO_NAME)

# Sep–Feb team (14)
TEAM_FEB = {
    'pilot': ["Aldrin Jhon Hina", "Arvin Anonuevo", "Dad Ivan Natividad", "Jay Ar Dastas"],
    'other': [
        "Christy Marasigan", "Jhan-Jhan Buac", "Mary Anne Melvida",
        "Ricardo Sabala Jr.", "Tricxie Mae Nazareta", "Jay Mark Toledana",
        "Ma. Cynthia Juano Mahinay", "Oscar Saverola Casimo",
        "Roselyn Espinosa Vargas", "Argenith Orlain Regonaos"
    ],
}
TEAM_FEB['all'] = TEAM_FEB['pilot'] + TEAM_FEB['other']

# Mar+ team (17)
TEAM_MAR = {
    'pilot': ["Aldrin Jhon Hina", "Dad Ivan Natividad", "Jay Ar Dastas"],
    'other': [
        "Arvin Anonuevo", "Christy Marasigan", "Jhan-Jhan Buac", "Mary Anne Melvida",
        "Ricardo Sabala Jr.", "Tricxie Mae Nazareta", "Jay Mark Toledana",
        "Ma. Cynthia Juano Mahinay", "Oscar Saverola Casimo",
        "Roselyn Espinosa Vargas", "Argenith Orlain Regonaos",
        "Henry Sabarillo", "Maria Sheena Garcia", "Robin Mark Quito"
    ],
}
TEAM_MAR['all'] = TEAM_MAR['pilot'] + TEAM_MAR['other']

# Combined set for any-month matching
ALL_ATS = sorted(set(TEAM_FEB['all'] + TEAM_MAR['all']))
ALL_ATS_SET = set(ALL_ATS)

# Month boundary: Mar starts at 2026-03
MAR_START = datetime(2026, 3, 1)

def get_team_for_month(month_label):
    """Return team config for a month label like 'Sep','Oct',...'Mar'"""
    mar_months = {'Mar', 'Apr', 'May'}
    if month_label in mar_months:
        return TEAM_MAR
    return TEAM_FEB

PLAN_FEB = {
    "Aldrin Jhon Hina": 1000000, "Arvin Anonuevo": 2194407.84,
    "Dad Ivan Natividad": 1218047, "Jay Ar Dastas": 1000000,
    "Christy Marasigan": 1369317, "Jhan-Jhan Buac": 1889766,
    "Mary Anne Melvida": 2577496, "Ricardo Sabala Jr.": 1464923,
    "Tricxie Mae Nazareta": 1000000, "Jay Mark Toledana": 1388050,
    "Ma. Cynthia Juano Mahinay": 1000000, "Oscar Saverola Casimo": 1000000,
    "Roselyn Espinosa Vargas": 1889169, "Argenith Orlain Regonaos": 1000000,
    "Henry Sabarillo": 1000000, "Maria Sheena Garcia": 1000000,
    "Robin Mark Quito": 1000000,
}
# Mar plan = 2x Feb actual (computed dynamically in generate_html)
PLAN_MAR = None  # will be set after parsing GR


def week_start(d):
    """Monday of the week containing d, with time stripped to midnight."""
    d0 = d.replace(hour=0, minute=0, second=0, microsecond=0)
    return d0 - timedelta(days=d0.weekday())


# ─── TERRITORY MAP ───────────────────────────────────────────────────
def load_territory_map(mgk_path, label=""):
    if not mgk_path.exists():
        print(f"  ⚠ Territory file not found: {mgk_path}")
        return {}, {}
    wb = openpyxl.load_workbook(mgk_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    terr_city, terr_brgy = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        prov = str(row[3] or '').strip().lower()
        city = str(row[4] or '').strip().lower()
        brgy = str(row[5] or '').strip().lower()
        raw_ats = str(row[6] or '').strip()
        ats = ATS_ALIASES.get(raw_ats, raw_ats)
        if ats and ats in ALL_ATS_SET and prov and city:
            terr_city[(prov, city)] = ats
            if brgy:
                terr_brgy[(prov, city, brgy)] = ats
    found = ALL_ATS_SET & (set(terr_city.values()) | set(terr_brgy.values()))
    print(f"  ✓ Territory {label}: {len(terr_city)} cities, {len(terr_brgy)} barangays, ATS coverage: {len(found)}")
    return terr_city, terr_brgy


# ─── PARSE GR (daily → weekly) ──────────────────────────────────────
def parse_gr(gr_path, terr_feb, terr_mar):
    """terr_feb = (terr_city, terr_brgy), terr_mar = (terr_city, terr_brgy)"""
    print(f"  Parsing: {gr_path.name}")
    wb = openpyxl.load_workbook(gr_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    dates = []
    for i in range(6, len(row2)):
        if not row2[i]: continue
        try:
            dates.append((i, datetime.strptime(str(row2[i]), '%m/%d/%Y')))
        except ValueError:
            pass

    all_weeks = sorted(set(week_start(d) for _, d in dates))
    NW = len(all_weeks)
    week_labels = [w.strftime("%b %d") for w in all_weeks]
    col_to_week = {ci: all_weeks.index(week_start(d)) for ci, d in dates}

    # Map each column to its territory set based on date
    col_to_terr = {}
    for ci, d in dates:
        if d >= MAR_START:
            col_to_terr[ci] = terr_mar
        else:
            col_to_terr[ci] = terr_feb

    month_map = {}
    for wi, w in enumerate(all_weeks):
        mk = w.strftime("%b")
        if mk not in month_map: month_map[mk] = []
        month_map[mk].append(wi)

    # Build per-month team info for JS
    team_by_month = {}
    for mk in month_map.keys():
        team = get_team_for_month(mk)
        team_by_month[mk] = {'pilot': team['pilot'], 'other': team['other'], 'all': team['all']}

    products_map = {'Applications scored': 's', 'Applications approved': 'a', 'COMBO Disbursement': 'd'}

    bm_w = {n: {'s': [0]*NW, 'a': [0]*NW, 'd': [0]*NW,
                '_sa': defaultdict(set), '_sl': defaultdict(set)} for n in ALL_ATS}

    current_ats_by_col = {}  # {col_index: ats_name} for current store
    current_store = None
    matched = unmatched = 0
    all_gr_stores = set()  # Collect all store names for matching engine
    gr_ats_stores = defaultdict(set)  # {ats: set of store_ids} for coverage calc

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            current_store = str(row[0]).strip()
            all_gr_stores.add(current_store)
            prov = str(row[1] or '').strip().lower()
            city = str(row[2] or '').strip().lower()
            brgy = str(row[3] or '').strip().lower()
            # Pre-compute ATS for each territory version
            tc_feb, tb_feb = terr_feb
            tc_mar, tb_mar = terr_mar
            ats_feb = tb_feb.get((prov, city, brgy)) or tc_feb.get((prov, city))
            ats_mar = tb_mar.get((prov, city, brgy)) or tc_mar.get((prov, city))
            if ats_feb: ats_feb = ATS_ALIASES.get(ats_feb, ats_feb)
            if ats_mar: ats_mar = ATS_ALIASES.get(ats_mar, ats_mar)
            # Collect store ID for coverage (use March territory as latest)
            store_m = re.match(r'^.*?\((\d+)\)\s*$', current_store)
            if store_m and ats_mar and ats_mar in ALL_ATS_SET:
                gr_ats_stores[ats_mar].add(store_m.group(1))
            current_ats_by_col = {}
            for ci, d in dates:
                ats = ats_mar if d >= MAR_START else ats_feb
                if ats and ats in ALL_ATS_SET:
                    current_ats_by_col[ci] = ats
            if current_ats_by_col:
                matched += 1
            else:
                unmatched += 1

        product = str(row[5] or '').strip()
        key = products_map.get(product)
        if not key or not current_ats_by_col: continue

        for ci, wi in col_to_week.items():
            ats = current_ats_by_col.get(ci)
            if not ats: continue
            val = row[ci] if ci < len(row) else None
            if val and isinstance(val, (int, float)):
                bm_w[ats][key][wi] += float(val)
                if key == 's' and val > 0: bm_w[ats]['_sa'][wi].add(current_store)
                if key == 'a' and val > 0: bm_w[ats]['_sl'][wi].add(current_store)

    for ats in ALL_ATS:
        bm_w[ats]['sa'] = [len(bm_w[ats]['_sa'].get(wi, set())) for wi in range(NW)]
        bm_w[ats]['sl'] = [len(bm_w[ats]['_sl'].get(wi, set())) for wi in range(NW)]
        del bm_w[ats]['_sa'], bm_w[ats]['_sl']
        bm_w[ats]['d'] = [round(v) for v in bm_w[ats]['d']]

    print(f"  ✓ {len(dates)} days → {NW} weeks, {len(month_map)} months | Stores: {matched} matched, {unmatched} non-NCR")
    # Compute max day of data per month from GR columns
    gr_max_day = {}  # {month_label: max_day_number}
    for ci, d in dates:
        mk = f"{d.year}-{d.month:02d}"
        for mkey, mi_list in month_map.items():
            wi = col_to_week.get(ci)
            if wi is not None and wi in mi_list:
                gr_max_day[mkey] = max(gr_max_day.get(mkey, 0), d.day)
                break

    for ats in ALL_ATS:
        td = sum(bm_w[ats]['d']); ts = sum(bm_w[ats]['s'])
        print(f"    {ats}: scored={ts:.0f}, disb=₱{td:,.0f}{' ⚠' if td==0 else ''}")

    return {'BM_W': bm_w, 'BM_WL': week_labels, 'BM_NW': NW,
            'BM_ML': list(month_map.keys()), 'BM_MR': list(month_map.values()),
            'TEAM_BY_MONTH': team_by_month, 'GR_MAX_DAY': gr_max_day,
            'GR_STORES': sorted(all_gr_stores),
            'GR_ATS_STORES': {ats: sorted(ids) for ats, ids in gr_ats_stores.items()}}


# ─── PARSE ACTIVATION STORES (New Stores tab) ───────────────────────
def parse_activation_stores(path):
    if not path.exists():
        print(f"  ⚠ Activation stores file not found: {path}")
        return None

    print(f"  Parsing: {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sup = str(row[6] or '').strip()
        if sup not in ALL_ATS_SET: continue
        stores.append({
            'sup': sup,
            'sr': row[7] if isinstance(row[7], datetime) else None,
            'fa': row[8] if isinstance(row[8], datetime) else None,
            'fl': row[9] if isinstance(row[9], datetime) else None,
            'status': str(row[10] or '').strip()
        })

    print(f"  ✓ {len(stores)} NCR stores loaded")

    # ── Weekly (last 8 weeks: Jan 5 → Feb 23 2026) ──
    # Determine recent weeks from data
    all_dates = [s['sr'] for s in stores if s['sr']] + [s['fa'] for s in stores if s['fa']] + [s['fl'] for s in stores if s['fl']]
    max_date = max(all_dates) if all_dates else datetime.now()

    # Last 8 weeks ending at max_date's week
    last_week = week_start(max_date)
    recent_weeks = [last_week - timedelta(weeks=i) for i in range(7, -1, -1)]
    rw_labels = [w.strftime("%b %-d") for w in recent_weeks]
    RW = len(recent_weeks)

    def aggregate_weekly(date_key):
        result = {n: [0]*RW for n in ALL_ATS}
        for s in stores:
            dt = s[date_key]
            if not dt: continue
            ws_d = week_start(dt)
            if ws_d in recent_weeks:
                wi = recent_weeks.index(ws_d)
                result[s['sup']][wi] += 1
        return result

    W_FL = aggregate_weekly('fl')
    W_FA = aggregate_weekly('fa')
    W_SR = aggregate_weekly('sr')

    # ── Monthly (Jan, Feb) ──
    month_labels_ns = ["Jan", "Feb"]

    def aggregate_monthly(date_key):
        result = {n: [0]*2 for n in ALL_ATS}
        for s in stores:
            dt = s[date_key]
            if not dt: continue
            if dt.year == 2026 and dt.month == 1: result[s['sup']][0] += 1
            elif dt.year == 2026 and dt.month == 2: result[s['sup']][1] += 1
        return result

    M_FL = aggregate_monthly('fl')
    M_FA = aggregate_monthly('fa')
    M_SR = aggregate_monthly('sr')

    # ── Overall (single bucket) ──
    def aggregate_overall(date_key):
        result = {n: [0] for n in ALL_ATS}
        for s in stores:
            dt = s[date_key]
            if not dt: continue
            if dt.year == 2026: result[s['sup']][0] += 1
        return result

    O_FL = aggregate_overall('fl')
    O_FA = aggregate_overall('fa')
    O_SR = aggregate_overall('sr')

    # ── Portfolio summary R ──
    R = {}
    for ats in ALL_ATS:
        ats_stores = [s for s in stores if s['sup'] == ats]
        # Get last week's BM data from stores (scored/approved/disb per week)
        # R needs: s, a, d arrays for the 8 recent weeks
        # But R in the dashboard actually holds weekly scored/approved/disb from GR
        # R is populated differently - it's the same as BM_W but filtered to recent weeks
        # Let's build the store portfolio stats instead
        total = len(ats_stores)
        active = sum(1 for s in ats_stores if s['status'] == '1st_loan_issued')
        app_only = sum(1 for s in ats_stores if s['status'] == '1st_application_done')
        approved = sum(1 for s in ats_stores if s['status'] == 'approved')
        new = sum(1 for s in ats_stores if s['status'] == 'new')
        R[ats] = {'t': total, 'active': active, 'app': app_only, 'approved': approved, 'new': new}

    print(f"  ✓ Weekly: {RW} weeks ({rw_labels[0]} → {rw_labels[-1]})")
    for ats in ALL_ATS:
        r = R[ats]
        wfl = sum(W_FL[ats]); wfa = sum(W_FA[ats]); wsr = sum(W_SR[ats])
        print(f"    {ats}: stores={r['t']}, active={r['active']} | new FL={wfl}, FA={wfa}, SR={wsr}")

    return {
        'WL': rw_labels, 'ML': month_labels_ns,
        'W_FL': W_FL, 'W_FA': W_FA, 'W_SR': W_SR,
        'M_FL': M_FL, 'M_FA': M_FA, 'M_SR': M_SR,
        'O_FL': O_FL, 'O_FA': O_FA, 'O_SR': O_SR,
        'R': R,
    }


# ─── GENERATE HTML ───────────────────────────────────────────────────
# ─── CHURN CALCULATION ────────────────────────────────────────────────
def calc_churn(gr_path, ns_path, terr_feb, terr_mar):
    """Calculate monthly churn rates from GR + activation stores.
    terr_feb/terr_mar = (terr_city, terr_brgy) tuples."""
    import re
    from collections import defaultdict
    tc_feb, tb_feb = terr_feb
    tc_mar, tb_mar = terr_mar

    # 1. Get stores with 1st_loan_issued (have first_loan_at date)
    wb_ns = openpyxl.load_workbook(ns_path, read_only=True)
    ws_ns = wb_ns[wb_ns.sheetnames[0]]
    loan_issued_ids = set()
    for row in ws_ns.iter_rows(min_row=2, values_only=True):
        store_id = str(row[0] or '').strip()
        fl_date = row[9]  # store_first_loan_at
        if fl_date and store_id:
            loan_issued_ids.add(store_id)

    # 2. Parse GR — get week dates and COMBO Disbursement per store
    wb_gr = openpyxl.load_workbook(gr_path, read_only=True)
    ws_gr = wb_gr[wb_gr.sheetnames[0]]
    row2 = list(ws_gr.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    week_dates = []
    for v in row2[6:]:
        if v:
            try:
                d = datetime.strptime(str(v), '%m/%d/%Y')
                week_dates.append(d)
            except:
                try:
                    d = datetime.strptime(str(v).split()[0], '%Y-%m-%d')
                    week_dates.append(d)
                except:
                    week_dates.append(None)
        else:
            week_dates.append(None)

    def week_month(d):
        return f"{d.year}-{d.month:02d}" if d else None

    def extract_id(name):
        m = re.search(r'\((\d+)\)', name)
        return m.group(1) if m else None

    store_ats = {}  # {store: {month_key: ats}}
    store_monthly_disb = defaultdict(lambda: defaultdict(float))
    current_store = ''
    store_prov = store_city = store_brgy = ''

    for row in ws_gr.iter_rows(min_row=3, values_only=True):
        if row[0]:
            current_store = str(row[0]).strip()
            store_prov = str(row[1] or '').strip().lower()
            store_city = str(row[2] or '').strip().lower()
            store_brgy = str(row[3] or '').strip().lower()
            # Pre-compute ATS for both territory sets
            ats_f = tb_feb.get((store_prov, store_city, store_brgy)) or tc_feb.get((store_prov, store_city))
            ats_m = tb_mar.get((store_prov, store_city, store_brgy)) or tc_mar.get((store_prov, store_city))
            if ats_f: ats_f = ATS_ALIASES.get(ats_f, ats_f)
            if ats_m: ats_m = ATS_ALIASES.get(ats_m, ats_m)
            store_ats[current_store] = {'feb': ats_f, 'mar': ats_m}

        metric = str(row[5] or '').strip()
        if metric == 'COMBO Disbursement' and current_store in store_ats:
            for i, v in enumerate(row[6:]):
                if i < len(week_dates) and week_dates[i]:
                    d = week_dates[i]
                    mk = week_month(d)
                    if mk and v and isinstance(v, (int, float)):
                        store_monthly_disb[current_store][mk] += v

    # Filter to qualified stores (1st_loan_issued)
    qualified = set()
    for sn in store_monthly_disb:
        sid = extract_id(sn)
        if sid and sid in loan_issued_ids:
            qualified.add(sn)

    all_months = sorted(set(m for sd in store_monthly_disb.values() for m in sd.keys()))

    # 3. Calculate churn per ATS per month
    month_name_map = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
                      '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
    churn_months = all_months[1:]  # skip first (base month)
    ch_labels = []
    for m in churn_months:
        yr = m[:4]
        mn = month_name_map[m[5:7]]
        ch_labels.append(f"{mn}-{yr[2:]}")

    def get_ats_for_month(sn, month_key):
        """Get ATS for a store in a given month (YYYY-MM)."""
        sa = store_ats.get(sn, {})
        if month_key >= '2026-03':
            ats = sa.get('mar')
        else:
            ats = sa.get('feb')
        return ats if ats and ats in ALL_ATS_SET else None

    ch_loan = {n: [] for n in ALL_ATS}
    for mi in range(1, len(all_months)):
        m_prev, m_curr = all_months[mi-1], all_months[mi]
        ats_active = defaultdict(set)
        ats_churned = defaultdict(set)
        for sn in qualified:
            ats = get_ats_for_month(sn, m_curr)
            if not ats:
                continue
            if store_monthly_disb[sn].get(m_prev, 0) > 0:
                ats_active[ats].add(sn)
                if store_monthly_disb[sn].get(m_curr, 0) == 0:
                    ats_churned[ats].add(sn)
        for n in ALL_ATS:
            ap = len(ats_active[n])
            ch = len(ats_churned[n])
            rate = round(ch / ap * 100, 1) if ap > 0 else 0
            ch_loan[n].append(rate)

    print(f"  ✓ Churn: {len(qualified)} stores, {len(churn_months)} months ({ch_labels[0]} → {ch_labels[-1]})")

    # ── Stores w Apps metrics ──
    # Parse Applications Scored per store per month
    store_monthly_apps = defaultdict(lambda: defaultdict(float))
    wb_gr2 = openpyxl.load_workbook(gr_path, read_only=True)
    ws_gr2 = wb_gr2[wb_gr2.sheetnames[0]]
    current_store2 = ''
    store_ats2 = {}  # {store: {feb: ats, mar: ats}}
    for row in ws_gr2.iter_rows(min_row=3, values_only=True):
        if row[0]:
            current_store2 = str(row[0]).strip()
            p2 = str(row[1] or '').strip().lower()
            c2 = str(row[2] or '').strip().lower()
            b2 = str(row[3] or '').strip().lower()
            af = tb_feb.get((p2, c2, b2)) or tc_feb.get((p2, c2))
            am = tb_mar.get((p2, c2, b2)) or tc_mar.get((p2, c2))
            if af: af = ATS_ALIASES.get(af, af)
            if am: am = ATS_ALIASES.get(am, am)
            if af or am:
                store_ats2[current_store2] = {'feb': af, 'mar': am}
        metric = str(row[5] or '').strip()
        if metric == 'Applications scored' and current_store2 in store_ats2:
            for i, v in enumerate(row[6:]):
                if i < len(week_dates) and week_dates[i]:
                    mk = week_month(week_dates[i])
                    if mk and v and isinstance(v, (int, float)):
                        store_monthly_apps[current_store2][mk] += v

    def get_ats2_for_month(sn, month_key):
        sa = store_ats2.get(sn, {})
        if month_key >= '2026-03':
            ats = sa.get('mar')
        else:
            ats = sa.get('feb')
        return ats if ats and ats in ALL_ATS_SET else None

    # Calculate per ATS per month: active, new, reactivated, churn (by apps)
    ch_apps_active = {n: [] for n in ALL_ATS}
    ch_apps_new = {n: [] for n in ALL_ATS}
    ch_apps_react = {n: [] for n in ALL_ATS}
    ch_apps_churn = {n: [] for n in ALL_ATS}

    for mi in range(1, len(all_months)):
        m_curr = all_months[mi]
        m_prev = all_months[mi-1]

        ats_active = defaultdict(int)
        ats_new = defaultdict(int)
        ats_react = defaultdict(int)
        ats_churn = defaultdict(int)

        for sn, apps in store_monthly_apps.items():
            ats = get_ats2_for_month(sn, m_curr)
            if not ats:
                continue

            curr = apps.get(m_curr, 0) > 0
            prev = apps.get(m_prev, 0) > 0
            ever_before = any(apps.get(all_months[j], 0) > 0 for j in range(mi))

            if curr:
                ats_active[ats] += 1
                if not ever_before:
                    ats_new[ats] += 1
                elif not prev and ever_before:
                    ats_react[ats] += 1

            if prev and not curr:
                ats_churn[ats] += 1

        for n in ALL_ATS:
            ch_apps_active[n].append(ats_active[n])
            ch_apps_new[n].append(ats_new[n])
            ch_apps_react[n].append(ats_react[n])
            ch_apps_churn[n].append(ats_churn[n])

    return {
        'CH_ML': ch_labels, 'CH_LOAN': ch_loan,
        'CH_APPS_ACTIVE': ch_apps_active, 'CH_APPS_NEW': ch_apps_new,
        'CH_APPS_REACT': ch_apps_react, 'CH_APPS_CHURN': ch_apps_churn,
    }


def generate_html(bm_data, ns_data, churn_data=None):
    with open(TEMPLATE_FILE) as f:
        html = f.read()

    wl = bm_data['BM_WL']
    date_range = f"{wl[0]} – {wl[-1]}" if wl else "No data"

    # Compute March plan = 2x Feb actual disbursement
    bm_w = bm_data['BM_W']
    bm_ml = bm_data['BM_ML']
    bm_mr = bm_data['BM_MR']
    plan_mar = {}
    new_mar_ats = {"Henry Sabarillo", "Maria Sheena Garcia", "Robin Mark Quito"}
    if 'Feb' in bm_ml:
        feb_idx = bm_ml.index('Feb')
        feb_weeks = bm_mr[feb_idx]
        for n in ALL_ATS:
            if n in new_mar_ats:
                plan_mar[n] = 1500000
            else:
                feb_disb = sum(bm_w[n]['d'][wi] for wi in feb_weeks)
                plan_mar[n] = round(feb_disb * 2)
    else:
        plan_mar = {n: 1000000 for n in ALL_ATS}

    # Build per-month PLAN: {month_label: {ats: plan}}
    plan_by_month = {}
    for mk in bm_ml:
        if mk in ('Mar', 'Apr', 'May'):
            plan_by_month[mk] = plan_mar
        else:
            plan_by_month[mk] = PLAN_FEB

    replacements = {
        '/*__BM_WL__*/': json.dumps(bm_data['BM_WL']),
        '/*__BM_NW__*/': str(bm_data['BM_NW']),
        '/*__BM_ML__*/': json.dumps(bm_data['BM_ML']),
        '/*__BM_MR__*/': json.dumps(bm_data['BM_MR']),
        '/*__BM_W__*/': json.dumps(bm_data['BM_W']),
        '/*__PLAN_BY_MONTH__*/': json.dumps(plan_by_month),
        '/*__TEAM_BY_MONTH__*/': json.dumps(bm_data.get('TEAM_BY_MONTH', {})),
        '/*__ALL_ATS__*/': json.dumps(ALL_ATS),
        '/*__GR_STORES__*/': json.dumps(bm_data.get('GR_STORES', [])),
        '/*__GR_ATS_STORES__*/': json.dumps(bm_data.get('GR_ATS_STORES', {})),
        '/*__UPDATE_DATE__*/': f"{date_range} (updated {datetime.now().strftime('%b %d, %H:%M')})",
    }

    if ns_data:
        # Build R in the format the dashboard expects:
        # R = {"ATS": {s:[...], a:[...], d:[...], sa:[...], sl:[...]}, ...}
        # These are the BM_W values for the last 8 weeks only
        bm_w = bm_data['BM_W']
        nw = bm_data['BM_NW']
        rw = min(8, nw)
        R_bm = {}
        for ats in ALL_ATS:
            R_bm[ats] = {
                's': bm_w[ats]['s'][-rw:],
                'a': bm_w[ats]['a'][-rw:],
                'd': bm_w[ats]['d'][-rw:],
                'sa': bm_w[ats]['sa'][-rw:],
                'sl': bm_w[ats]['sl'][-rw:],
            }

        replacements.update({
            '/*__NS_WL__*/': json.dumps(ns_data['WL']),
            '/*__NS_ML__*/': json.dumps(ns_data['ML']),
            '/*__NS_W_FL__*/': json.dumps(ns_data['W_FL']),
            '/*__NS_W_FA__*/': json.dumps(ns_data['W_FA']),
            '/*__NS_W_SR__*/': json.dumps(ns_data['W_SR']),
            '/*__NS_M_FL__*/': json.dumps(ns_data['M_FL']),
            '/*__NS_M_FA__*/': json.dumps(ns_data['M_FA']),
            '/*__NS_M_SR__*/': json.dumps(ns_data['M_SR']),
            '/*__NS_O_FL__*/': json.dumps(ns_data['O_FL']),
            '/*__NS_O_FA__*/': json.dumps(ns_data['O_FA']),
            '/*__NS_O_SR__*/': json.dumps(ns_data['O_SR']),
            '/*__NS_R__*/': json.dumps(R_bm),
        })

    # Churn data
    if churn_data:
        replacements['/*__CH_ML__*/'] = json.dumps(churn_data['CH_ML'])
        replacements['/*__CH_LOAN__*/'] = json.dumps(churn_data['CH_LOAN'])
        replacements['/*__CH_APPS_ACTIVE__*/'] = json.dumps(churn_data['CH_APPS_ACTIVE'])
        replacements['/*__CH_APPS_NEW__*/'] = json.dumps(churn_data['CH_APPS_NEW'])
        replacements['/*__CH_APPS_REACT__*/'] = json.dumps(churn_data['CH_APPS_REACT'])
        replacements['/*__CH_APPS_CHURN__*/'] = json.dumps(churn_data['CH_APPS_CHURN'])
    else:
        replacements['/*__CH_ML__*/'] = '[]'
        replacements['/*__CH_LOAN__*/'] = '{}'
        replacements['/*__CH_APPS_ACTIVE__*/'] = '{}'
        replacements['/*__CH_APPS_NEW__*/'] = '{}'
        replacements['/*__CH_APPS_REACT__*/'] = '{}'
        replacements['/*__CH_APPS_CHURN__*/'] = '{}'

    # Generate MONTH_DAYS based on actual GR data dates (not current date)
    bm_months = bm_data['BM_ML']
    gr_max_day = bm_data.get('GR_MAX_DAY', {})
    month_days = {}
    month_map_num = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
    for m in bm_months:
        mn = month_map_num.get(m, 1)
        yr = 2025 if mn >= 6 else 2026
        import calendar
        total = calendar.monthrange(yr, mn)[1]
        elapsed = gr_max_day.get(m, 0)
        if elapsed == 0:
            # Fallback: if month is fully past, set elapsed = total
            today = datetime.now()
            if (yr < today.year) or (yr == today.year and mn < today.month):
                elapsed = total
        month_days[m] = {'total': total, 'elapsed': elapsed}

    replacements['/*__MONTH_DAYS__*/'] = json.dumps(month_days)

    for marker, value in replacements.items():
        if marker not in html:
            print(f"  ⚠ Marker not found: {marker}")
        html = html.replace(marker, value)
    return html


# ─── GIT DEPLOY ──────────────────────────────────────────────────────
def git_deploy():
    os.chdir(REPO_DIR)
    try:
        subprocess.run(["git", "add", "index.html"], check=True, capture_output=True)
        msg = f"Dashboard update {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        result = subprocess.run(["git", "commit", "-m", msg], capture_output=True, text=True)
        if "nothing to commit" in result.stdout:
            print("  ℹ No changes to deploy")
            return False
        subprocess.run(["git", "push"], check=True, capture_output=True)
        print("  ✓ Pushed → GitHub Pages will auto-deploy")
        return True
    except subprocess.CalledProcessError as e:
        print(f"  ✗ Git error: {e.stderr}")
        return False


# ─── MAIN ────────────────────────────────────────────────────────────
def main():
    print("╔══════════════════════════════════════╗")
    print("║  NCR Pilot Dashboard — Auto-updater  ║")
    print("╚══════════════════════════════════════╝")
    print(f"\nData dir: {DATA_DIR}")
    print(f"Output:   {OUTPUT_FILE}\n")

    if not DATA_DIR.exists():
        print(f"✗ Data directory not found: {DATA_DIR}"); sys.exit(1)
    if not TEMPLATE_FILE.exists():
        print(f"✗ Template not found: {TEMPLATE_FILE}"); sys.exit(1)

    # 1. Territory maps (Feb + Mar)
    print("[1/5] Loading territory maps...")

    # Auto-detect territory files: mgk_ath*.xlsx or *territory*.xlsx
    def find_territory(pattern_list):
        for pat in pattern_list:
            hits = sorted(DATA_DIR.glob(pat))
            if hits:
                return hits[-1]
        return None

    # Feb territory: mgk_ath.xlsx or *territory*02*.xlsx or *territory*feb*.xlsx
    feb_terr_path = find_territory([
        "mgk_ath.xlsx",
        "mgk_ath_territory_split_03_02*.xlsx",
        "*territory*02*copy*.xlsx",
        "*territory*02*.xlsx",
        "*territory*feb*.xlsx",
    ])
    if feb_terr_path:
        terr_feb = load_territory_map(feb_terr_path, "Feb")
    else:
        print("  ✗ No Feb territory file found!"); sys.exit(1)

    # Mar territory: mgk_ath_mar.xlsx or *territory*03_03*.xlsx
    mar_terr_path = find_territory([
        "mgk_ath_mar.xlsx",
        "mgk_ath_territory_split_03_03*.xlsx",
        "*territory*03_03*.xlsx",
        "*territory*mar*.xlsx",
    ])
    if mar_terr_path:
        terr_mar = load_territory_map(mar_terr_path, "Mar")
    else:
        print("  ⚠ No March territory file, using Feb for all months")
        terr_mar = terr_feb

    # 2. Parse GR
    print("[2/5] Parsing General Report...")
    gr_path = None
    for name in ["GR_daily.xlsx", "GR_weekly.xlsx"]:
        p = DATA_DIR / name
        if p.exists():
            gr_path = p; break
    if not gr_path:
        gr_files = sorted(DATA_DIR.glob("General_Report*.xlsx")) + sorted(DATA_DIR.glob("GR*.xlsx"))
        gr_path = gr_files[-1] if gr_files else None
    if not gr_path or not gr_path.exists():
        print("  ✗ No GR file found!"); sys.exit(1)
    bm_data = parse_gr(gr_path, terr_feb, terr_mar)

    # 3. Parse Activation Stores
    print("[3/5] Parsing Activation Stores...")
    ns_path = None
    for name in ["activation_stores.xlsx", "Overall_status_of_activation_stores.xlsx"]:
        p = DATA_DIR / name
        if p.exists():
            ns_path = p; break
    if not ns_path:
        ns_files = sorted(DATA_DIR.glob("*activation*.xlsx")) + sorted(DATA_DIR.glob("*Overall*.xlsx"))
        ns_path = ns_files[0] if ns_files else None
    ns_data = parse_activation_stores(ns_path) if ns_path and ns_path.exists() else None

    # 3b. Calculate Churn
    churn_data = calc_churn(gr_path, ns_path, terr_feb, terr_mar) if ns_path else None

    # 4. Generate HTML
    print("[4/5] Generating dashboard HTML...")
    html = generate_html(bm_data, ns_data, churn_data)
    with open(OUTPUT_FILE, 'w') as f:
        f.write(html)
    print(f"  ✓ Written: {OUTPUT_FILE.name} ({len(html):,} bytes)")

    # 5. Deploy
    print("[5/5] Deploying...")
    git_deploy()
    print("\n✅ Done!")


if __name__ == "__main__":
    main()
